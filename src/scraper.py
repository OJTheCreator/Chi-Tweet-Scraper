"""
Twitter Scraper - Clean Working Version
========================================
Based on the original working code with minimal fixes for:
1. Stopping properly when date range is complete
2. Better error handling
"""

import os
import json
import asyncio
import csv
import random
from datetime import datetime, timezone
from twikit import Client, TooManyRequests
from openpyxl import Workbook, load_workbook
import logging
import pandas as pd
import re
from collections import defaultdict
import traceback

TWEET_ID_PATTERN = re.compile(
    r"^https?://(?:www\.)?(?:twitter\.com|x\.com)/\w+/status/(\d+)"
)

# ========================================
# CONFIGURATION
# ========================================
RATE_LIMIT_DELAY = 3
MAX_NETWORK_RETRIES = 10
RETRY_DELAYS = [10, 30, 60, 120, 300, 600, 900]
MAX_PAGINATION_RETRIES = 10
MAX_CONSECUTIVE_EMPTY_PAGES = 150
EMPTY_PAGE_PROMPT_THRESHOLD = 25
MAX_CURSOR_REFRESHES = 10
MAX_CONSECUTIVE_ERRORS = 20
AUTO_SAVE_INTERVAL = 25

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
COOKIES_FILE = os.path.join(BASE_DIR, "cookies", "twikit_cookies.json")
DEFAULT_EXPORT_DIR = os.path.join(BASE_DIR, "data", "exports")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ========================================
# EXCEPTION CLASSES
# ========================================
class TwitterScraperError(Exception):
    pass


class CookieExpiredError(TwitterScraperError):
    pass


class NetworkError(TwitterScraperError):
    pass


class RateLimitError(TwitterScraperError):
    pass


class EmptyPagePromptException(TwitterScraperError):
    pass


class ScraperPausedException(TwitterScraperError):
    pass


class TwitterAPIError(TwitterScraperError):
    pass


class CookieConflictError(TwitterScraperError):
    """Raised when there are duplicate cookies that need to be cleaned."""

    pass


# ========================================
# ERROR CLASSIFICATION FUNCTIONS
# ========================================
def is_cookie_conflict_error(error_msg: str) -> bool:
    """Check if error is due to duplicate cookies."""
    return "multiple cookies exist with name" in error_msg.lower()


def is_network_error(error_msg: str) -> bool:
    error_lower = error_msg.lower()
    keywords = [
        "connection",
        "timeout",
        "network",
        "unreachable",
        "timed out",
        "connection reset",
        "connection refused",
        "temporary failure",
        "getaddrinfo",
        "11001",
        "errno 11001",
        "name resolution",
        "ssl",
        "certificate",
        "handshake",
        "eof",
        "broken pipe",
        "connection aborted",
        "remote end closed",
        "socket",
        "dns",
        "host",
        "resolve",
        "econnreset",
        "econnrefused",
        "etimedout",
        "enetunreach",
        "ehostunreach",
        "epipe",
        "no route to host",
        "network is down",
        "temporarily unavailable",
    ]
    return any(k in error_lower for k in keywords)


def is_auth_error(error_msg: str) -> bool:
    error_lower = error_msg.lower()
    keywords = [
        "unauthorized",
        "forbidden",
        "authentication",
        "token",
        "expired",
        "401",
        "403",
        "login",
        "credential",
        "session",
        "invalid cookie",
        "cookie expired",
        "not authenticated",
    ]
    return any(k in error_lower for k in keywords)


def is_rate_limit_error(error_msg: str) -> bool:
    error_lower = error_msg.lower()
    keywords = [
        "rate limit",
        "too many requests",
        "429",
        "slow down",
        "try again later",
        "exceeded",
        "throttle",
    ]
    return any(k in error_lower for k in keywords)


def is_twitter_api_error(error_msg: str) -> bool:
    error_lower = error_msg.lower()
    keywords = [
        "twitter",
        "api",
        "500",
        "502",
        "503",
        "504",
        "internal server",
        "service unavailable",
        "bad gateway",
        "over capacity",
    ]
    return any(k in error_lower for k in keywords)


# ========================================
# UTILITY FUNCTIONS
# ========================================
def clean_duplicate_cookies(cookies_file: str) -> bool:
    """Remove duplicate cookies from the cookie file."""
    try:
        if not os.path.exists(cookies_file):
            return False

        with open(cookies_file, "r", encoding="utf-8") as f:
            cookies = json.load(f)

        if not isinstance(cookies, list):
            return False

        seen = {}
        for cookie in cookies:
            if isinstance(cookie, dict):
                name = cookie.get("name", "")
                domain = cookie.get("domain", "")
                key = (name, domain)
                seen[key] = cookie

        cleaned_cookies = list(seen.values())

        if len(cleaned_cookies) < len(cookies):
            with open(cookies_file, "w", encoding="utf-8") as f:
                json.dump(cleaned_cookies, f, indent=2)
            logger.info(
                f"Cleaned {len(cookies) - len(cleaned_cookies)} duplicate cookies"
            )
            return True

        return False
    except Exception as e:
        logger.warning(f"Failed to clean cookies: {e}")
        return False


async def smart_sleep(
    seconds: int, should_stop_callback=None, progress_callback=None, message_prefix=""
):
    """Sleep with periodic progress updates and cancellation support."""
    for remaining in range(seconds, 0, -1):
        if should_stop_callback and should_stop_callback():
            raise asyncio.CancelledError("Stopped during wait")
        await asyncio.sleep(1)
        if remaining % 30 == 0 and remaining > 0 and progress_callback:
            minutes, secs = divmod(remaining, 60)
            progress_callback(f"{message_prefix}Resuming in {minutes:02d}:{secs:02d}")


# ========================================
# AUTHENTICATION
# ========================================
async def authenticate(retry_callback=None, should_stop_callback=None):
    """Authenticate with automatic cookie deduplication."""

    if os.path.exists(COOKIES_FILE):
        if clean_duplicate_cookies(COOKIES_FILE):
            if retry_callback:
                retry_callback("🧹 Cleaned duplicate cookies from file")

    for attempt in range(MAX_NETWORK_RETRIES):
        if should_stop_callback and should_stop_callback():
            raise asyncio.CancelledError("Authentication stopped by user")

        try:
            client = Client(language="en-US")
            if not os.path.exists(COOKIES_FILE):
                raise CookieExpiredError(
                    "Cookie file not found. Please save cookies first."
                )

            client.load_cookies(COOKIES_FILE)

            if should_stop_callback and should_stop_callback():
                raise asyncio.CancelledError("Authentication stopped by user")

            await client.search_tweet("(from:twitter)", product="Latest")
            logger.info("Authentication successful")
            return client

        except asyncio.CancelledError:
            raise
        except CookieExpiredError:
            raise
        except Exception as e:
            if should_stop_callback and should_stop_callback():
                raise asyncio.CancelledError("Authentication stopped by user")

            error_msg = str(e)

            if is_cookie_conflict_error(error_msg):
                if retry_callback:
                    retry_callback("🧹 Fixing duplicate cookie issue...")
                clean_duplicate_cookies(COOKIES_FILE)
                continue

            if is_auth_error(error_msg):
                raise CookieExpiredError(
                    "Authentication failed. Cookies may be expired."
                )

            if is_network_error(error_msg):
                if attempt < MAX_NETWORK_RETRIES - 1:
                    delay = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)]
                    if retry_callback:
                        retry_callback(
                            f"🔌 Network error. Retrying in {delay}s... ({attempt + 1}/{MAX_NETWORK_RETRIES})"
                        )
                    await smart_sleep(delay, should_stop_callback)
                    continue
                raise NetworkError(
                    f"Network failed after {MAX_NETWORK_RETRIES} attempts."
                )

            if attempt < MAX_NETWORK_RETRIES - 1:
                delay = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)]
                if retry_callback:
                    retry_callback(f"⚠️ Auth error: {str(e)[:50]}. Retrying...")
                await smart_sleep(delay, should_stop_callback)
                continue

            raise TwitterScraperError(f"Authentication failed: {e}")

    raise NetworkError("Failed to authenticate after maximum retries")


async def handle_network_retry(
    operation,
    progress_callback=None,
    cookie_expired_callback=None,
    should_stop_callback=None,
    max_retries=None,
):
    """Execute an operation with comprehensive retry logic."""
    max_retries = max_retries or MAX_NETWORK_RETRIES
    cookie_cleaned = False

    for attempt in range(max_retries):
        if should_stop_callback and should_stop_callback():
            raise asyncio.CancelledError("Operation stopped by user")
        try:
            return await operation()
        except asyncio.CancelledError:
            raise
        except CookieExpiredError:
            if cookie_expired_callback:
                cookie_expired_callback("Cookie expired")
            raise
        except TooManyRequests:
            if progress_callback:
                progress_callback("⏳ Rate limit hit. Waiting 15 minutes...")
            await smart_sleep(
                900, should_stop_callback, progress_callback, "⏳ Rate limit: "
            )
            continue
        except Exception as e:
            error_msg = str(e)

            if is_cookie_conflict_error(error_msg) and not cookie_cleaned:
                if progress_callback:
                    progress_callback("🧹 Fixing duplicate cookie issue...")
                clean_duplicate_cookies(COOKIES_FILE)
                cookie_cleaned = True
                continue

            if is_auth_error(error_msg):
                if cookie_expired_callback:
                    cookie_expired_callback(error_msg)
                raise CookieExpiredError(error_msg)

            if is_rate_limit_error(error_msg):
                if progress_callback:
                    progress_callback("⏳ Rate limit. Waiting 15 minutes...")
                await smart_sleep(
                    900, should_stop_callback, progress_callback, "⏳ Rate limit: "
                )
                continue

            if is_network_error(error_msg) or is_twitter_api_error(error_msg):
                if attempt < max_retries - 1:
                    delay = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)]
                    if progress_callback:
                        progress_callback(
                            f"🔌 Error. Retrying in {delay}s... ({attempt + 1}/{max_retries})"
                        )
                    await smart_sleep(delay, should_stop_callback)
                    continue

            if attempt < max_retries - 1:
                delay = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)]
                if progress_callback:
                    progress_callback(f"⚠️ Error: {str(e)[:60]}. Retrying...")
                await smart_sleep(delay, should_stop_callback)
                continue

            raise TwitterScraperError(f"Operation failed: {e}")

    raise TwitterScraperError("Operation failed after maximum retries")


# ========================================
# QUERY AND DATA FUNCTIONS
# ========================================
def validate_date_range(start_date: str, end_date: str) -> tuple:
    try:
        start_dt = None
        end_dt = None
        if start_date:
            date_part = start_date.split("_")[0] if "_" in start_date else start_date
            start_dt = datetime.strptime(date_part, "%Y-%m-%d")
        if end_date:
            date_part = end_date.split("_")[0] if "_" in end_date else end_date
            end_dt = datetime.strptime(date_part, "%Y-%m-%d")
        today = datetime.now()
        if end_dt and end_dt > today:
            end_dt = today
        if start_dt and end_dt and start_dt > end_dt:
            raise TwitterScraperError("Start date cannot be after end date.")
        return start_dt, end_dt
    except ValueError as e:
        raise TwitterScraperError(f"Invalid date format: {e}")


def build_search_query(
    username=None, keywords=None, start_date=None, end_date=None, use_and=False
):
    if not (username or keywords):
        raise TwitterScraperError("Either username or keywords must be provided.")
    if username:
        username = username.strip().lstrip("@")
        query = f"(from:{username}) -filter:replies"
    else:
        clean_keywords = [kw.strip() for kw in keywords if kw.strip()]
        if not clean_keywords:
            raise TwitterScraperError("No valid keywords provided.")
        operator = " AND " if use_and else " OR "
        keyword_query = operator.join([f'"{kw}"' for kw in clean_keywords])
        query = f"({keyword_query}) -filter:replies"
    if start_date:
        query += (
            f" since:{start_date.split('_')[0] if '_' in start_date else start_date}"
        )
    if end_date:
        query += f" until:{end_date.split('_')[0] if '_' in end_date else end_date}"
    return query


def extract_tweet_data(tweet) -> dict:
    """Extract tweet data with robust error handling."""
    try:
        created_at = getattr(tweet, "created_at", "") or getattr(
            tweet, "created_at_datetime", ""
        )
        try:
            if created_at:
                if isinstance(created_at, str):
                    dt = datetime.strptime(created_at, "%a %b %d %H:%M:%S %z %Y")
                else:
                    dt = created_at
                formatted_date = dt.strftime("%Y-%m-%d %H:%M:%S")
            else:
                formatted_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        except:
            formatted_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        user = getattr(tweet, "user", None) or getattr(tweet, "author", None)
        username = ""
        display_name = ""
        if user:
            username = getattr(user, "screen_name", "") or getattr(user, "username", "")
            display_name = getattr(user, "name", "")

        text = getattr(tweet, "text", "") or getattr(tweet, "full_text", "")
        if text:
            text = text.replace("\n", " ").replace("\r", " ")

        tweet_id = getattr(tweet, "id", "") or getattr(tweet, "id_str", "")

        views = 0
        try:
            if hasattr(tweet, "view_count") and tweet.view_count:
                views = int(tweet.view_count)
            elif hasattr(tweet, "views") and tweet.views:
                views = (
                    int(tweet.views.get("count", 0))
                    if isinstance(tweet.views, dict)
                    else int(tweet.views)
                )
        except:
            pass

        data = {
            "date": formatted_date,
            "username": username,
            "display_name": display_name,
            "text": text,
            "retweets": getattr(tweet, "retweet_count", 0) or 0,
            "likes": getattr(tweet, "favorite_count", 0)
            or getattr(tweet, "like_count", 0)
            or 0,
            "replies": getattr(tweet, "reply_count", 0) or 0,
            "quotes": getattr(tweet, "quote_count", 0) or 0,
            "views": views,
            "tweet_id": tweet_id,
            "tweet_url": (
                f"https://twitter.com/{username}/status/{tweet_id}"
                if username and tweet_id
                else ""
            ),
        }

        if not data["tweet_id"] or not data["text"]:
            return None
        return data
    except Exception as e:
        logger.warning(f"Error extracting tweet data: {e}")
        return None


def sanitize_worksheet_name(name: str) -> str:
    from unicodedata import normalize

    name = normalize("NFKD", name)
    emoji_pattern = re.compile(
        "[\U0001f600-\U0001f64f\U0001f300-\U0001f5ff\U0001f680-\U0001f6ff"
        "\U0001f1e0-\U0001f1ff\U00002702-\U000027b0\U000024c2-\U0001f251"
        "\U0001f900-\U0001f9ff\U0001fa70-\U0001faff]+",
        flags=re.UNICODE,
    )
    name = emoji_pattern.sub("", name)
    for char in [
        "\\",
        "/",
        "*",
        "[",
        "]",
        ":",
        "?",
        "|",
        "(",
        ")",
        "<",
        ">",
        '"',
        "'",
        "{",
        "}",
    ]:
        name = name.replace(char, "_")
    name = re.sub(r"[^A-Za-z0-9\s_\-]", "", name)
    name = re.sub(r"[\s_\-]+", "_", name).strip("_ -")
    if len(name) > 31:
        name = name[:31].rstrip("_-")
    return name if name else f"Sheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}"


def should_include_tweet(tweet_data, keywords=None, use_and=False):
    if not keywords:
        return True
    text_lower = tweet_data["text"].lower()
    matches = [k.lower() in text_lower for k in keywords]
    return all(matches) if use_and else any(matches)


async def take_custom_break(
    break_settings, current_count, progress_callback=None, should_stop_callback=None
):
    if not break_settings or not break_settings.get("enabled"):
        return
    interval = break_settings.get("tweet_interval", 100)
    if current_count > 0 and current_count % interval == 0:
        mins = random.randint(
            break_settings.get("min_break_minutes", 5),
            break_settings.get("max_break_minutes", 10),
        )
        if progress_callback:
            progress_callback(f"☕ Taking {mins}-min break... ({current_count} tweets)")
        await smart_sleep(
            mins * 60, should_stop_callback, progress_callback, "☕ Break: "
        )
        if progress_callback:
            progress_callback("🔄 Break complete! Resuming...")


# ========================================
# MAIN SCRAPING FUNCTION
# ========================================
async def scrape_tweets(
    username=None,
    start_date=None,
    end_date=None,
    keywords=None,
    use_and=False,
    export_format="excel",
    save_dir=DEFAULT_EXPORT_DIR,
    progress_callback=None,
    should_stop_callback=None,
    cookie_expired_callback=None,
    network_error_callback=None,
    save_every_n=AUTO_SAVE_INTERVAL,
    max_tweets=None,
    break_settings=None,
    resume_state=None,
):
    """
    Scrape tweets - Clean working version.
    Based on original code that worked, with proper stop conditions.
    """
    csv_file = None
    wb = None
    ws = None
    writer = None

    # Track why we stopped - ALWAYS log this at the end
    stop_reason = "Unknown"

    try:
        # Clean cookies at start
        clean_duplicate_cookies(COOKIES_FILE)

        resuming = resume_state is not None
        if resuming:
            if progress_callback:
                progress_callback("🔄 Resuming from previous session...")
            count = resume_state.get("tweets_scraped", 0)
            seen_tweet_ids = set(resume_state.get("seen_tweet_ids", []))
            output_path = resume_state.get("output_path")
            if export_format.lower() == "csv":
                csv_file = open(output_path, mode="a", newline="", encoding="utf-8")
                writer = csv.writer(csv_file)
            else:
                wb = load_workbook(output_path)
                ws = wb.active
        else:
            validate_date_range(start_date, end_date)
            count = 0
            seen_tweet_ids = set()
            os.makedirs(save_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "csv" if export_format.lower() == "csv" else "xlsx"
            output_path = os.path.join(
                save_dir, f"{username or 'keywords'}_{timestamp}.{ext}"
            )
            headers = [
                "Date",
                "Username",
                "Display Name",
                "Text",
                "Retweets",
                "Likes",
                "Replies",
                "Quotes",
                "Views",
                "Tweet ID",
                "Tweet URL",
                "Export Path",
            ]
            if export_format.lower() == "csv":
                csv_file = open(output_path, mode="w", newline="", encoding="utf-8")
                writer = csv.writer(csv_file)
                writer.writerow(headers)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sanitize_worksheet_name(
                    username or "_".join((keywords or [])[:3]) or "Tweets"
                )
                ws.append(headers)

        query = build_search_query(username, keywords, start_date, end_date, use_and)
        if progress_callback and not resuming:
            progress_callback(f"🔍 Query: {query}")
        if progress_callback:
            progress_callback("🔑 Authenticating...")

        client = await authenticate(
            retry_callback=progress_callback, should_stop_callback=should_stop_callback
        )

        if should_stop_callback and should_stop_callback():
            return output_path, count, list(seen_tweet_ids)

        if progress_callback:
            progress_callback("🔍 Starting search...")

        # Parse dates for comparison
        start_dt = (
            datetime.strptime(start_date.split("_")[0], "%Y-%m-%d")
            if start_date
            else None
        )
        end_dt = (
            datetime.strptime(end_date.split("_")[0], "%Y-%m-%d") if end_date else None
        )

        oldest_tweet_date = None
        newest_tweet_date = None
        reached_start_date = False

        async def start_search():
            return await client.search_tweet(query, product="Latest")

        page = await handle_network_retry(
            start_search,
            progress_callback,
            cookie_expired_callback,
            should_stop_callback,
        )

        empty_page_count = 0
        cursor_refresh_count = 0
        consecutive_errors = 0
        total_pages = 0
        last_save = 0
        cookie_clean_attempts = 0  # Track cookie cleaning attempts

        try:
            while page:
                if should_stop_callback and should_stop_callback():
                    stop_reason = "User cancelled"
                    break

                total_pages += 1
                page_tweets = 0
                page_oldest_date = None

                # Process tweets from page
                for tweet in page:
                    if should_stop_callback and should_stop_callback():
                        stop_reason = "User cancelled during processing"
                        break
                    if max_tweets and count >= max_tweets:
                        stop_reason = f"Reached max tweets limit ({max_tweets})"
                        break

                    tweet_data = extract_tweet_data(tweet)
                    if not tweet_data:
                        continue

                    tid = tweet_data["tweet_id"]
                    if tid in seen_tweet_ids:
                        continue
                    seen_tweet_ids.add(tid)

                    if not should_include_tweet(tweet_data, keywords, use_and):
                        continue

                    # Track dates
                    try:
                        td = datetime.strptime(tweet_data["date"], "%Y-%m-%d %H:%M:%S")

                        if oldest_tweet_date is None or td < oldest_tweet_date:
                            oldest_tweet_date = td
                        if newest_tweet_date is None or td > newest_tweet_date:
                            newest_tweet_date = td
                        if page_oldest_date is None or td < page_oldest_date:
                            page_oldest_date = td

                        # Check if this tweet is AT or BEFORE start date
                        if start_dt and td.date() <= start_dt.date():
                            reached_start_date = True
                    except:
                        pass

                    # Save tweet
                    row = [
                        tweet_data["date"],
                        tweet_data["username"],
                        tweet_data["display_name"],
                        tweet_data["text"],
                        tweet_data["retweets"],
                        tweet_data["likes"],
                        tweet_data["replies"],
                        tweet_data["quotes"],
                        tweet_data["views"],
                        tweet_data["tweet_id"],
                        tweet_data["tweet_url"],
                        os.path.abspath(output_path),
                    ]

                    if export_format.lower() == "csv":
                        writer.writerow(row)
                    else:
                        ws.append(row)

                    count += 1
                    page_tweets += 1
                    consecutive_errors = 0

                    if progress_callback:
                        progress_callback(count)

                    # Auto-save
                    if count - last_save >= save_every_n:
                        if export_format.lower() == "csv":
                            csv_file.flush()
                        else:
                            wb.save(output_path)
                        last_save = count
                        if progress_callback:
                            progress_callback(f"💾 Auto-saved {count} tweets")

                    await take_custom_break(
                        break_settings, count, progress_callback, should_stop_callback
                    )

                # Handle empty pages
                if page_tweets == 0:
                    empty_page_count += 1

                    if progress_callback and empty_page_count % 5 == 0:
                        progress_callback(
                            f"📭 Empty page {empty_page_count} ({count} tweets so far)"
                        )

                    # FIX: If we've reached start date, stop after some empty pages
                    if reached_start_date and empty_page_count >= 10:
                        stop_reason = f"Date range complete (reached start date + {empty_page_count} empty pages)"
                        if progress_callback:
                            progress_callback(
                                f"✅ Date range complete! {count} tweets collected."
                            )
                        break

                    # Only stop if we have NO tweets at all after many attempts
                    if count == 0 and empty_page_count >= 20:
                        stop_reason = "No tweets found after 20 empty pages"
                        if progress_callback:
                            progress_callback("❌ No tweets found matching criteria")
                        break
                else:
                    empty_page_count = 0

                # Check if we've passed start date
                if page_oldest_date and start_dt:
                    if page_oldest_date.date() < start_dt.date():
                        stop_reason = f"Successfully reached start date (oldest: {page_oldest_date.strftime('%Y-%m-%d')})"
                        if progress_callback:
                            progress_callback(
                                f"✅ Reached tweets before start date ({page_oldest_date.strftime('%Y-%m-%d')}). "
                                f"{count} tweets collected."
                            )
                        reached_start_date = True
                        break

                # ========================================
                # PAGINATION
                # ========================================
                pag_success = False
                pag_attempt = 0

                while not pag_success and pag_attempt < MAX_PAGINATION_RETRIES:
                    try:
                        if should_stop_callback and should_stop_callback():
                            raise asyncio.CancelledError()

                        if progress_callback:
                            progress_callback(
                                f"📄 Loading next page... ({count} tweets)"
                            )

                        page = await page.next()
                        pag_success = True
                        consecutive_errors = 0

                    except StopAsyncIteration:
                        # End of pagination - attempt cursor refresh if needed
                        if not reached_start_date and start_dt and oldest_tweet_date:
                            days_left = (
                                oldest_tweet_date.date() - start_dt.date()
                            ).days

                            if (
                                days_left > 0
                                and cursor_refresh_count < MAX_CURSOR_REFRESHES
                            ):
                                if progress_callback:
                                    progress_callback(
                                        f"⚠️ ~{days_left} days remaining. "
                                        f"Refreshing cursor ({cursor_refresh_count + 1}/{MAX_CURSOR_REFRESHES})..."
                                    )

                                try:
                                    await asyncio.sleep(5)
                                    refresh_until = oldest_tweet_date.strftime(
                                        "%Y-%m-%d"
                                    )
                                    refresh_query = build_search_query(
                                        username,
                                        keywords,
                                        start_date,
                                        refresh_until,
                                        use_and,
                                    )

                                    if progress_callback:
                                        progress_callback(
                                            f"🔄 New search: {start_date.split('_')[0] if start_date else 'N/A'} to {refresh_until}"
                                        )

                                    page = await client.search_tweet(
                                        refresh_query, product="Latest"
                                    )
                                    cursor_refresh_count += 1
                                    empty_page_count = 0
                                    pag_success = True

                                    if progress_callback:
                                        progress_callback("✅ Cursor refreshed!")
                                    continue
                                except Exception as re:
                                    if progress_callback:
                                        progress_callback(
                                            f"⚠️ Refresh failed: {str(re)[:50]}"
                                        )

                            # Failed to get more data
                            stop_reason = f"End of pagination with ~{days_left} days remaining (tried {cursor_refresh_count} refreshes)"
                        else:
                            if reached_start_date:
                                stop_reason = "Successfully completed date range"
                            else:
                                stop_reason = "End of available results"

                        if progress_callback:
                            progress_callback(
                                f"📋 End of available results. {count} tweets collected."
                            )
                        page = None
                        break

                    except TooManyRequests:
                        if progress_callback:
                            progress_callback("⏳ Rate limit. Waiting 15 min...")
                        await smart_sleep(
                            900, should_stop_callback, progress_callback, "⏳ "
                        )
                        pag_attempt = 0
                        continue

                    except asyncio.CancelledError:
                        raise

                    except Exception as e:
                        em = str(e)
                        consecutive_errors += 1
                        pag_attempt += 1

                        if is_cookie_conflict_error(em):
                            cookie_clean_attempts += 1

                            if cookie_clean_attempts <= 3:
                                if progress_callback:
                                    progress_callback(
                                        f"🧹 Fixing duplicate cookie issue (attempt {cookie_clean_attempts}/3)..."
                                    )
                                clean_duplicate_cookies(COOKIES_FILE)
                                # Try to reload cookies in the client
                                try:
                                    client.load_cookies(COOKIES_FILE)
                                    pag_attempt = 0
                                    continue
                                except:
                                    pass

                            # After 3 attempts, prompt for new cookies
                            if progress_callback:
                                progress_callback(
                                    "🔑 Cookie issue persists. Please provide new cookies."
                                )
                            if cookie_expired_callback:
                                cookie_expired_callback(
                                    f"Cookie error after {cookie_clean_attempts} cleaning attempts: {em}"
                                )
                            raise CookieExpiredError(f"Cookie error persists: {em}")

                        if is_auth_error(em):
                            if cookie_expired_callback:
                                cookie_expired_callback("Session expired")
                            raise CookieExpiredError("Session expired")

                        if is_rate_limit_error(em):
                            if progress_callback:
                                progress_callback("⏳ Rate limit...")
                            await smart_sleep(
                                900, should_stop_callback, progress_callback, "⏳ "
                            )
                            pag_attempt = 0
                            continue

                        if is_network_error(em) or is_twitter_api_error(em):
                            if pag_attempt >= MAX_PAGINATION_RETRIES:
                                # Exhausted retries - raise exception to trigger GUI dialog
                                stop_reason = f"Network error after {pag_attempt} retries: {em[:50]}"
                                if network_error_callback:
                                    network_error_callback(em)
                                raise NetworkError(
                                    f"Network error persists after {pag_attempt} retries: {em}"
                                )

                            delay = RETRY_DELAYS[
                                min(pag_attempt - 1, len(RETRY_DELAYS) - 1)
                            ]
                            if progress_callback:
                                progress_callback(
                                    f"🔌 Error. Waiting {delay}s... ({pag_attempt}/{MAX_PAGINATION_RETRIES})"
                                )
                            await smart_sleep(delay, should_stop_callback)
                            continue

                        if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                            stop_reason = f"Too many consecutive errors ({consecutive_errors}): {em[:50]}"
                            # Raise exception to trigger GUI dialog
                            raise NetworkError(
                                f"Too many errors ({consecutive_errors}): {em}"
                            )

                        delay = min(30 * pag_attempt, 300)
                        if progress_callback:
                            progress_callback(
                                f"⚠️ Error: {str(e)[:60]}. Retrying in {delay}s..."
                            )
                        await smart_sleep(delay, should_stop_callback)
                        continue

                if page is None:
                    break

        except asyncio.CancelledError:
            stop_reason = "Cancelled by user"
            if progress_callback:
                progress_callback("🛑 Cancelled")
            raise
        finally:
            if export_format.lower() == "csv" and csv_file:
                csv_file.close()
                csv_file = None
            elif wb:
                wb.save(output_path)

        # Final verification
        if start_dt and oldest_tweet_date:
            if oldest_tweet_date.date() > start_dt.date():
                days_miss = (oldest_tweet_date.date() - start_dt.date()).days
                if stop_reason == "Unknown":
                    stop_reason = f"Incomplete - missing ~{days_miss} days"
                if progress_callback:
                    progress_callback(
                        f"⚠️ Could not reach full date range. Missing ~{days_miss} days. "
                        f"Oldest tweet: {oldest_tweet_date.strftime('%Y-%m-%d')}"
                    )
            else:
                if stop_reason == "Unknown":
                    stop_reason = "Successfully completed date range"
                if progress_callback:
                    progress_callback(
                        f"✅ Date range fully covered: {oldest_tweet_date.strftime('%Y-%m-%d')} to "
                        f"{newest_tweet_date.strftime('%Y-%m-%d') if newest_tweet_date else 'N/A'}"
                    )

        # Always log the stop reason
        if progress_callback:
            progress_callback(f"📝 Stop reason: {stop_reason}")
            progress_callback(f"✅ Complete: {count} tweets ({total_pages} pages)")

        return output_path, count, list(seen_tweet_ids)

    except CookieExpiredError:
        if progress_callback:
            progress_callback("🔑 Cookie expired")
        raise
    except NetworkError:
        if progress_callback:
            progress_callback("🔌 Network error")
        raise
    except asyncio.CancelledError:
        raise
    except Exception as e:
        logger.error(f"Scrape error: {e}\n{traceback.format_exc()}")
        raise TwitterScraperError(f"Scraping failed: {e}")
    finally:
        if csv_file and not csv_file.closed:
            csv_file.close()


# ========================================
# LINK SCRAPING
# ========================================
async def scrape_tweet_links_file(
    file_path,
    export_format="excel",
    save_dir=DEFAULT_EXPORT_DIR,
    progress_callback=None,
    should_stop_callback=None,
    cookie_expired_callback=None,
    network_error_callback=None,
    break_settings=None,
    resume_state=None,
):
    """Scrape tweets from a file of links."""
    csv_file = None
    wb = None
    ws = None
    writer = None
    client = None

    try:
        clean_duplicate_cookies(COOKIES_FILE)

        resuming = resume_state is not None
        if resuming:
            if progress_callback:
                progress_callback("🔄 Resuming link scrape...")
            file_path = resume_state.get("links_file_path")
            output_path = resume_state.get("output_path")
            count = resume_state.get("tweets_scraped", 0)
            failed = resume_state.get("failed_count", 0)
            skipped = resume_state.get("skipped_no_data", 0)
            current_index = resume_state.get("current_index", 0)
            processed_links = set(resume_state.get("processed_links", []))
            if export_format.lower() == "csv":
                csv_file = open(output_path, mode="a", newline="", encoding="utf-8")
                writer = csv.writer(csv_file)
            else:
                wb = load_workbook(output_path)
                ws = wb.active
        else:
            if not os.path.exists(file_path):
                raise TwitterScraperError(f"File not found: {file_path}")
            count = failed = skipped = current_index = 0
            processed_links = set()
            os.makedirs(save_dir, exist_ok=True)
            ext = "csv" if export_format.lower() == "csv" else "xlsx"
            output_path = os.path.join(
                save_dir,
                f"tweet_links_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}",
            )
            headers = [
                "Date",
                "Username",
                "Display Name",
                "Text",
                "Retweets",
                "Likes",
                "Replies",
                "Quotes",
                "Views",
                "Tweet ID",
                "Tweet URL",
                "Export Path",
            ]
            if export_format.lower() == "csv":
                csv_file = open(output_path, mode="w", newline="", encoding="utf-8")
                writer = csv.writer(csv_file)
                writer.writerow(headers)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Tweets"
                ws.append(headers)

        # Load links
        ext = os.path.splitext(file_path)[1].lower()
        if ext in [".xlsx", ".xls"]:
            df = pd.read_excel(file_path, header=None)
            links = df.iloc[:, 0].dropna().astype(str).tolist()
        elif ext == ".txt":
            with open(file_path, "r", encoding="utf-8") as f:
                links = [l.strip() for l in f if l.strip()]
        else:
            raise TwitterScraperError("Use .txt or .xlsx/.xls")

        valid_links = [
            l for l in links if TWEET_ID_PATTERN.match(l) and l not in processed_links
        ]
        if not valid_links:
            raise TwitterScraperError("No valid links found")

        if progress_callback:
            progress_callback("🔑 Authenticating...")
        client = await authenticate(
            retry_callback=progress_callback, should_stop_callback=should_stop_callback
        )
        total = len(valid_links) + current_index

        for i, link in enumerate(valid_links, current_index + 1):
            if should_stop_callback and should_stop_callback():
                break
            if progress_callback:
                progress_callback(f"🔗 Link {i}/{total}")

            retries = 0
            while retries < 5:
                try:
                    match = TWEET_ID_PATTERN.match(link)
                    if not match:
                        failed += 1
                        break
                    tweet = await client.get_tweet_by_id(match.group(1))
                    td = extract_tweet_data(tweet)
                    if not td:
                        skipped += 1
                        processed_links.add(link)
                        break
                    row = [
                        td["date"],
                        td["username"],
                        td["display_name"],
                        td["text"],
                        td["retweets"],
                        td["likes"],
                        td["replies"],
                        td["quotes"],
                        td["views"],
                        td["tweet_id"],
                        td["tweet_url"],
                        os.path.abspath(output_path),
                    ]
                    if export_format.lower() == "csv":
                        writer.writerow(row)
                    else:
                        ws.append(row)
                    count += 1
                    processed_links.add(link)
                    if progress_callback:
                        progress_callback(f"✅ {count} scraped")
                    if count % 20 == 0:
                        if export_format.lower() == "csv":
                            csv_file.flush()
                        else:
                            wb.save(output_path)
                    await asyncio.sleep(RATE_LIMIT_DELAY)
                    await take_custom_break(
                        break_settings, count, progress_callback, should_stop_callback
                    )
                    break
                except Exception as e:
                    em = str(e)

                    if is_cookie_conflict_error(em):
                        if progress_callback:
                            progress_callback("🧹 Fixing cookie issue...")
                        clean_duplicate_cookies(COOKIES_FILE)
                        retries += 1
                        continue

                    if is_auth_error(em):
                        raise CookieExpiredError("Session expired")
                    if is_network_error(em):
                        retries += 1
                        if retries < 5:
                            delay = RETRY_DELAYS[
                                min(retries - 1, len(RETRY_DELAYS) - 1)
                            ]
                            if progress_callback:
                                progress_callback(f"🔌 Retrying in {delay}s...")
                            await asyncio.sleep(delay)
                            continue
                        if network_error_callback:
                            network_error_callback(em)
                        raise NetworkError(f"Network failed: {e}")
                    failed += 1
                    processed_links.add(link)
                    await asyncio.sleep(RATE_LIMIT_DELAY)
                    break

        if export_format.lower() == "csv" and csv_file:
            csv_file.close()
            csv_file = None
        elif wb:
            wb.save(output_path)

        if progress_callback:
            progress_callback(
                f"🏁 Done: {count} scraped, {failed} failed, {skipped} skipped"
            )
        return output_path, count, failed, list(processed_links)

    except CookieExpiredError:
        if progress_callback:
            progress_callback("🔑 Cookie expired")
        raise
    except NetworkError:
        if progress_callback:
            progress_callback("🔌 Network error")
        raise
    except asyncio.CancelledError:
        raise
    except Exception as e:
        logger.error(f"Link scrape error: {e}")
        raise TwitterScraperError(f"Link scraping failed: {e}")
    finally:
        if csv_file and not csv_file.closed:
            csv_file.close()
        if client and hasattr(client, "close"):
            try:
                await client.close()
            except:
                pass


# ========================================
# MULTI-USER SCRAPING
# ========================================
async def scrape_multiple_usernames(
    usernames,
    start_date,
    end_date,
    export_format="excel",
    save_dir=DEFAULT_EXPORT_DIR,
    progress_callback=None,
    should_stop_callback=None,
    cookie_expired_callback=None,
    max_tweets_per_user=None,
    break_settings=None,
):
    if not usernames:
        raise TwitterScraperError("No usernames provided")
    results = []
    total = 0
    for i, u in enumerate(usernames, 1):
        if should_stop_callback and should_stop_callback():
            break
        clean = u.strip().lstrip("@")
        if not clean:
            continue
        if progress_callback:
            progress_callback(f"👤 User {i}/{len(usernames)}: @{clean}")
        try:
            out, cnt, _ = await scrape_tweets(
                username=clean,
                start_date=start_date,
                end_date=end_date,
                export_format=export_format,
                save_dir=save_dir,
                progress_callback=progress_callback,
                should_stop_callback=should_stop_callback,
                cookie_expired_callback=cookie_expired_callback,
                max_tweets=max_tweets_per_user,
                break_settings=break_settings,
            )
            results.append(
                {
                    "username": clean,
                    "output_path": out,
                    "tweet_count": cnt,
                    "status": "success",
                }
            )
            total += cnt
            if progress_callback:
                progress_callback(f"✅ @{clean}: {cnt} tweets")
        except CookieExpiredError:
            raise
        except NetworkError:
            raise
        except Exception as e:
            results.append(
                {
                    "username": clean,
                    "output_path": None,
                    "tweet_count": 0,
                    "status": "failed",
                    "error": str(e),
                }
            )
            if progress_callback:
                progress_callback(f"❌ @{clean} failed")
    if progress_callback:
        ok = len([r for r in results if r["status"] == "success"])
        progress_callback(f"🎉 Batch: {ok}/{len(usernames)} users, {total} tweets")
    return results
